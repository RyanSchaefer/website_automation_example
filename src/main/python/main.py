from fbs_runtime.application_context.PyQt5 import ApplicationContext
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import sys
from PyQt5.QtWidgets import QFileDialog
from selenium.webdriver.chrome.options import Options
chrome_options = Options()
chrome_options.add_argument("--log-level=3")
import urllib
urllib.urlretrieve('ftp://server/path/to/file', 'file')

try:
    wb = openpyxl.load_workbook("input.xlsx")
    sheet = wb["input"]
except:
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = "output"

appctxt = ApplicationContext()

b = webdriver.Chrome(appctxt.get_resource("chromedriver.exe"), chrome_options=chrome_options)
b.get(r"file://" + appctxt.get_resource("example-site.html"))


# some elements do not require any special procedures
input = b.find_element_by_id("input")
input.send_keys("some search")

# some elements will require wrappers for proper functionality
select = Select(b.find_element_by_id("select"))

# on some websites there will not be an id for elements so we have to go by XPath
button = b.find_element_by_xpath("/html/body/table[1]/tbody/tr[3]/td[2]/button")

for x in range(3):
    select.select_by_index(x)
    button.click()

input.clear()
input.send_keys("another search")

for x in range(3):
    select.select_by_index(x)
    button.click()

input.clear()
input.send_keys("final search")

for x in range(3):
    select.select_by_index(2-x)
    button.click()

table = b.find_element_by_id("output_table")

for x, row in enumerate(table.find_elements_by_tag_name("tr")):
    for y, cell in enumerate(row.find_elements_by_tag_name("td")):
        wb["output"].cell(row=x+1, column=y+1).value = cell.text



if __name__ == '__main__':
    file_dialog = QFileDialog()
    file = str(file_dialog.getExistingDirectory(caption="Save to Folder", directory="C:\\"))
    print(file)
    wb.save(file + r"/output.xlsx")
    b.quit()
    # exit_code = appctxt.app.exec_()      # 2. Invoke appctxt.app.exec_()
    sys.exit(0)